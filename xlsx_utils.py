"""
Microsoft Excel XLSX file utilities.

@author: Nik Mohamad Aizuddin bin Nik Azmi
@email: nik-mohamad-aizuddin@yandex.com
@version: 0.1.0
@date: 23/May/2018
"""

import csv

import numpy as np
from openpyxl import load_workbook

def read(fname_xlsx):
    """Load features and ideal output from XLSX file.

    Parameters
    ----------
    fname_xlsx : str
        Excel XLSX filename to read.

    Returns
    -------
    features : np.ndarray
        Features with shape (n,3).
    ideal_output : np.ndarray
        Ideal output with shape (n,2).
    header : list
        XLSX table header.
    """

    # Load Sheet1 from excel file.
    workbook = load_workbook(filename=fname_xlsx)
    sheet = workbook['Sheet1']

    # Extract table header.
    header = []
    columns = sheet.max_column
    for i in sheet.iter_cols(min_col=1, max_col=columns):
        header.append(i[0].value)

    # Extract number of room, room size, and corridor as features.
    # Also extract the value limit and colony as the ideal output.
    features = []
    ideal_output = []
    
    #change here for different dataset read 
    rows = sheet.max_row   
    for i in sheet.iter_rows(min_row=2, max_row=rows):   
        features.append([i[k].value for k in range(0, (columns-2))])
        ideal_output.append([i[k].value for k in range((columns-2), columns)])

    # Convert from list to numpy array.
    features = np.asarray(features).astype(np.float)
    ideal_output = np.asarray(ideal_output).astype(np.float)

    return features, ideal_output, header


def read_train(fname_xlsx):
    """Load features and ideal output from XLSX file.

    Parameters
    ----------
    fname_xlsx : str
        Excel XLSX filename to read.

    Returns
    -------
    features : np.ndarray
        Features with shape (n,3).
    ideal_output : np.ndarray
        Ideal output with shape (n,2).
    header : list
        XLSX table header.
    """

    # Load Sheet1 from excel file.
    workbook = load_workbook(filename=fname_xlsx)
    sheet = workbook['Sheet1']

    # Extract table header.
    header = []
    columns = sheet.max_column  
    for i in sheet.iter_cols(min_col=1, max_col=columns):
        header.append(i[0].value)

    # Extract number of room, room size, and corridor as features.
    # Also extract the value limit and colony as the ideal output.
    features = []
    ideal_output = []
    
    #change here for different dataset read 
    rows = sheet.max_row
    for i in sheet.iter_rows(min_row=2, max_row=rows):   
        features.append([i[k].value for k in range(0, (columns-2))])
        ideal_output.append([i[k].value for k in range((columns-2), columns)])

    # Convert from list to numpy array.
    features = np.asarray(features).astype(np.float)
    ideal_output = np.asarray(ideal_output).astype(np.float)

    return features, ideal_output, header

def read_test(fname_xlsx):
    """Load features and ideal output from XLSX file.

    Parameters
    ----------
    fname_xlsx : str
        Excel XLSX filename to read.

    Returns
    -------
    features : np.ndarray
        Features with shape (n,3).
    ideal_output : np.ndarray
        Ideal output with shape (n,2).
    header : list
        XLSX table header.
    """

    # Load Sheet1 from excel file.
    workbook = load_workbook(filename=fname_xlsx)
    sheet = workbook['Sheet1']

    # Extract table header.
    header = []
    columns = sheet.max_column 
    for i in sheet.iter_cols(min_col=1, max_col=columns):
        header.append(i[0].value)

    # Extract number of room, room size, and corridor as features.
    # Also extract the value limit and colony as the ideal output.
    features = []
    ideal_output = []
    
    #change here for different dataset read 
    rows = sheet.max_row
    for i in sheet.iter_rows(min_row=2, max_row=rows):   
        features.append([i[k].value for k in range(0, (columns-2))])
        ideal_output.append([i[k].value for k in range((columns-2), columns)])

    # Convert from list to numpy array.
    features = np.asarray(features).astype(np.float)
    ideal_output = np.asarray(ideal_output).astype(np.float)

    return features, ideal_output, header